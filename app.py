from datetime import datetime
import io
import os
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="Control de Compromisos y Producción", layout="wide"
)

st.title("📊 Panel de Control: Seguimiento Comercial")
st.markdown("Consolidación inteligente con formatos contables (Soles, Porcentajes y Enteros) y control de historial.")

st.sidebar.header("📂 Gestión de Archivos")

# 1. Subir el archivo consolidado base
archivo_base = st.sidebar.file_uploader(
    "1. Sube tu Excel Consolidado o Modelo Base", type=["xlsx"], key="base"
)

# 2. Subir los reportes individuales semanales de los vendedores
archivos_nuevos = st.sidebar.file_uploader(
    "2. Sube los reportes individuales semanales de los vendedores",
    type=["xlsx"],
    accept_multiple_files=True,
    key="nuevos",
)

if "df_auditoria" not in st.session_state:
  st.session_state["df_auditoria"] = pd.DataFrame()
if "wb_template" not in st.session_state:
  st.session_state["wb_template"] = None

# Cargar plantilla base
if archivo_base is not None:
  try:
    archivo_base.seek(0)
    st.session_state["wb_template"] = openpyxl.load_workbook(archivo_base)

    archivo_base.seek(0)
    xls_base = pd.ExcelFile(archivo_base)

    if "Detalle_Auditoria" in xls_base.sheet_names:
      df_temp = pd.read_excel(xls_base, sheet_name="Detalle_Auditoria", header=0)
      df_temp.columns = [str(c).strip() for c in df_temp.columns]
      df_temp = df_temp.dropna(subset=["Cliente"]) if "Cliente" in df_temp.columns else df_temp
      st.session_state["df_auditoria"] = df_temp

    st.sidebar.success("✅ ¡Plantilla base cargada con éxito!")
  except Exception as e:
    st.sidebar.error(f"⚠️ Error al leer el archivo base: {e}")

nombre_archivo_modelo_github = (
    "Modelo - Control Compromisos para Ventas (FECHA).xlsx"
)
if os.path.exists(nombre_archivo_modelo_github):
  with open(nombre_archivo_modelo_github, "rb") as f:
    bytes_modelo = f.read()
  st.sidebar.download_button(
      label="📥 Descargar Formato Oficial de Vendedor",
      data=bytes_modelo,
      file_name="Formato_Oficial_Vendedor.xlsx",
      mime=(
          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
      ),
  )

st.sidebar.markdown("---")

# --- PROCESAMIENTO Y VALIDACIÓN DE REPORTES ---
if archivos_nuevos:
  lista_nuevos_dfs = []
  
  columnas_obligatorias = [
      "Cliente", "No. Contacto", "1o. Contacto", "Ultimo Contacto", 
      "N° Cotización", "Unidad", "Cantidad", "Valor + IGV", 
      "Chance de Venta", "Mes Previsto"
  ]

  for file in archivos_nuevos:
    try:
      file.seek(0)
      df_v = pd.read_excel(file, sheet_name=0, header=0)
      df_v.columns = [str(c).strip() for c in df_v.columns]

      faltantes = [col for col in columnas_obligatorias if col not in df_v.columns]
      tiene_responsable = "CODIGO-RESPONSABLE" in df_v.columns or "RESPONSABLE" in df_v.columns
      if not tiene_responsable:
        faltantes.append("CODIGO-RESPONSABLE (o RESPONSABLE)")

      if faltantes:
        st.sidebar.error(
            f"❌ **Rechazado ('{file.name}')**: Le falta(n) la(s) columna(s): **{', '.join(faltantes)}**."
        )
        continue

      df_v = df_v.dropna(subset=["Cliente"])

      df_normalizado = pd.DataFrame()
      df_normalizado["Cliente"] = df_v["Cliente"]
      df_normalizado["No. Contacto"] = df_v["No. Contacto"]
      df_normalizado["1o. Contacto"] = df_v["1o. Contacto"]
      df_normalizado["Ultimo Contacto"] = df_v["Ultimo Contacto"]
      df_normalizado["N° Cotización"] = df_v["N° Cotización"]
      df_normalizado["Unidad"] = df_v["Unidad"]
      
      df_normalizado["Cantidad"] = pd.to_numeric(df_v["Cantidad"], errors="coerce").fillna(0).round(0).astype(int)
      df_normalizado["Valor + IGV"] = pd.to_numeric(df_v["Valor + IGV"], errors="coerce").fillna(0)
      
      # --- AGREGADO: Status Resumido ---
      if "Status Resumido" in df_v.columns:
          df_normalizado["Status Resumido"] = df_v["Status Resumido"]
      else:
          df_normalizado["Status Resumido"] = ""
      
      df_normalizado["Chance de Venta"] = pd.to_numeric(df_v["Chance de Venta"], errors="coerce").fillna(0)
      
      df_normalizado["Valor Ponderado (S/)"] = df_normalizado["Valor + IGV"] * df_normalizado["Chance de Venta"]
      df_normalizado["Cantidad Ponderada Prod."] = df_normalizado["Cantidad"] * df_normalizado["Chance de Venta"]
      
      df_normalizado["Mes Previsto"] = df_v["Mes Previsto"]
      
      resp_col = "CODIGO-RESPONSABLE" if "CODIGO-RESPONSABLE" in df_v.columns else "RESPONSABLE"
      df_normalizado["RESPONSABLE"] = df_v[resp_col]

      lista_nuevos_dfs.append(df_normalizado)
      st.sidebar.success(f"✅ ¡Vendedor validado y procesado: {file.name}!")
    except Exception as e:
      st.sidebar.error(f"❌ Error al procesar '{file.name}': {e}")

  if lista_nuevos_dfs:
    df_agregado = pd.concat(lista_nuevos_dfs, ignore_index=True)
    
    if not st.session_state["df_auditoria"].empty:
      df_combinado = pd.concat([st.session_state["df_auditoria"], df_agregado], ignore_index=True)
    else:
      df_combinado = df_agregado

    df_combinado = df_combinado.drop_duplicates(
        subset=["N° Cotización", "Ultimo Contacto", "Chance de Venta"], 
        keep="last"
    )
    
    st.session_state["df_auditoria"] = df_combinado

# --- GENERACIÓN DEL EXCEL FINAL ---
if not st.session_state["df_auditoria"].empty:
  st.sidebar.markdown("---")
  fecha_actual = datetime.now().strftime("%d-%m-%y")
  nombre_archivo = f"Control_Compromisos_Ventas_Final_{fecha_actual}.xlsx"

  output = io.BytesIO()
  if st.session_state["wb_template"] is not None:
    wb = st.session_state["wb_template"]
    if "Detalle_Auditoria" in wb.sheetnames:
      ws = wb["Detalle_Auditoria"]
      if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
      
      df_final = st.session_state["df_auditoria"].dropna(subset=["Cliente"])
      registros = df_final.to_dict("records")
      
      for r_idx, row in enumerate(registros, start=2):
        ws.cell(row=r_idx, column=1, value=row.get("Cliente", ""))
        ws.cell(row=r_idx, column=2, value=row.get("No. Contacto", ""))
        ws.cell(row=r_idx, column=3, value=row.get("1o. Contacto", ""))
        ws.cell(row=r_idx, column=4, value=row.get("Ultimo Contacto", ""))
        ws.cell(row=r_idx, column=5, value=row.get("N° Cotización", ""))
        ws.cell(row=r_idx, column=6, value=row.get("Unidad", ""))
        
        cant = int(pd.to_numeric(row.get("Cantidad", 0), errors="coerce") or 0)
        importe = float(pd.to_numeric(row.get("Valor + IGV", 0), errors="coerce") or 0)
        chance = float(pd.to_numeric(row.get("Chance de Venta", 0), errors="coerce") or 0)
        
        ws.cell(row=r_idx, column=7, value=cant).number_format = '#,##0'
        ws.cell(row=r_idx, column=8, value=importe).number_format = 'S/ #,##0.00'
        ws.cell(row=r_idx, column=9, value=chance).number_format = '0.00%'
        
        ws.cell(row=r_idx, column=10, value=importe * chance).number_format = 'S/ #,##0.00'
        ws.cell(row=r_idx, column=11, value=cant * chance).number_format = '#,##0.00'
        
        ws.cell(row=r_idx, column=12, value=row.get("Mes Previsto", ""))
        ws.cell(row=r_idx, column=13, value=row.get("RESPONSABLE", ""))

      data_font = Font(name="Calibri", size=10)
      border_style = Border(
          left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
          top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
      )

      for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
          cell.font = data_font
          cell.border = border_style
          cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(output)
    excel_data = output.getvalue()
  else:
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
      st.session_state["df_auditoria"].to_excel(writer, sheet_name="Detalle_Auditoria", index=False)
    excel_data = output.getvalue()

  st.sidebar.download_button(
      label="📥 Descargar Consolidado Final Alineado",
      data=excel_data,
      file_name=nombre_archivo,
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
  )

# --- VISUALIZACIÓN EN PESTAÑAS CON BUSCADORES Y FILTROS ---
tab1, tab2, tab3 = st.tabs([
    "📦 Resumen por Mes y Unidad",
    "💰 Seguimiento Ventas",
    "📋 Detalle de Auditoría",
])

with tab1:
  st.subheader("Resumen Proyectado: Cantidades y Valores por Unidad (Chance > 0%)")
  if not st.session_state["df_auditoria"].empty:
    df_aud = st.session_state["df_auditoria"].copy()
    df_aud["Cantidad_Num"] = pd.to_numeric(df_aud["Cantidad"], errors="coerce").fillna(0).astype(int)
    df_aud["Importe_Num"] = pd.to_numeric(df_aud["Valor + IGV"], errors="coerce").fillna(0)
    df_aud["Chance Num"] = pd.to_numeric(df_aud["Chance de Venta"], errors="coerce").fillna(0)
    df_aud = df_aud[df_aud["Chance Num"] > 0]
    
    if not df_aud.empty and "Mes Previsto" in df_aud.columns and "Unidad" in df_aud.columns:
      def formatear_mes_yyyy_mm(val):
        if pd.isna(val): return "Sin Especificar"
        try:
          return pd.to_datetime(val).strftime("%Y-%m")
        except:
          s = str(val).strip()
          return s[:7] if len(s) >= 7 else s

      df_aud["Mes Formateado"] = df_aud["Mes Previsto"].apply(formatear_mes_yyyy_mm)
      meses_disponibles = sorted([str(m) for m in df_aud["Mes Formateado"].unique()])
      mes_actual_default = datetime.now().strftime("%Y-%m")
      defaults = [mes_actual_default] if mes_actual_default in meses_disponibles else meses_disponibles
      
      col_f1, col_f2 = st.columns(2)
      with col_f1:
        meses_seleccionados = st.multiselect("Filtrar Mes Previsto:", options=meses_disponibles, default=defaults, key="m_tab1")
      with col_f2:
        buscar_tab1 = st.text_input("🔍 Buscar en Resumen (Unidad o Mes):", "", key="b_tab1")

      if meses_seleccionados:
        df_aud = df_aud[df_aud["Mes Formateado"].isin(meses_seleccionados)]

      df_resumen = df_aud.groupby(["Mes Formateado", "Unidad"], dropna=False).agg(
          Total_Cantidad=("Cantidad_Num", "sum"),
          Total_Valor_IGV=("Importe_Num", "sum"),
          Total_Cotizaciones=("N° Cotización", "count")
      ).reset_index()

      if buscar_tab1:
        mask = df_resumen.astype(str).apply(lambda x: x.str.contains(buscar_tab1, case=False, na=False)).any(axis=1)
        df_resumen = df_resumen[mask]
      
      df_resumen_display = df_resumen.copy()
      df_resumen_display["Total_Cantidad"] = df_resumen_display["Total_Cantidad"].apply(lambda x: f"{x:,.0f}")
      df_resumen_display["Total_Valor_IGV"] = df_resumen_display["Total_Valor_IGV"].apply(lambda x: f"S/ {x:,.2f}")
      df_resumen_display["Total_Cotizaciones"] = df_resumen_display["Total_Cotizaciones"].astype(str)
      
      st.dataframe(df_resumen_display.style.set_properties(**{'text-align': 'left'}), use_container_width=True)
      
      total_general_cant = df_resumen["Total_Cantidad"].sum() if not df_resumen.empty else 0
      total_general_val = df_resumen["Total_Valor_IGV"].sum() if not df_resumen.empty else 0
      
      col1, col2 = st.columns(2)
      with col1:
        st.metric(label="📦 Cantidad Total Filtrada", value=f"{int(total_general_cant):,}")
      with col2:
        st.metric(label="💰 Valor Total con IGV (S/)", value=f"S/ {total_general_val:,.2f}")
    else:
      st.warning("No hay registros con Chance de Venta mayor a 0%.")
  else:
    st.info("Sube tus archivos para visualizar el resumen.")

with tab2:
  st.subheader("Seguimiento de Ventas (Calculado Automáticamente)")
  if not st.session_state["df_auditoria"].empty:
    df_ventas_calc = st.session_state["df_auditoria"].copy()
    
    # Asegurar que las columnas sean numéricas para poder sumarlas
    df_ventas_calc["Importe_Num"] = pd.to_numeric(df_ventas_calc["Valor + IGV"], errors="coerce").fillna(0)
    df_ventas_calc["Cantidad_Num"] = pd.to_numeric(df_ventas_calc["Cantidad"], errors="coerce").fillna(0).astype(int)
    
    if "Mes Previsto" in df_ventas_calc.columns and "Unidad" in df_ventas_calc.columns:
      df_ventas_calc["Mes Formateado"] = df_ventas_calc["Mes Previsto"].apply(
          lambda x: pd.to_datetime(x).strftime("%Y-%m") if pd.notna(x) and not pd.isna(pd.to_datetime(x, errors='coerce')) else str(x)[:7]
      )
      
      # Agrupar sumando la Cantidad y eliminando el Ponderado
      df_seg_base = df_ventas_calc.groupby(["Mes Formateado", "Unidad"], dropna=False).agg(
          Cantidad=("Cantidad_Num", "sum"),
          N_Cotizaciones=("N° Cotización", "count"),
          Pipeline_Bruto=("Importe_Num", "sum")
      ).reset_index()

      # Ordenar las columnas para que "Cantidad" quede al lado de "Unidad"
      df_seg_base = df_seg_base[["Mes Formateado", "Unidad", "Cantidad", "N_Cotizaciones", "Pipeline_Bruto"]]

      # Estandarizar el nombre de la unidad para facilitar el filtrado
      df_seg_base['Unidad_lower'] = df_seg_base['Unidad'].astype(str).str.lower().str.strip()

      # --- 1. TABLA BOLSAS ---
      df_bolsas = df_seg_base[df_seg_base['Unidad_lower'] == 'bolsas'].drop(columns=['Unidad_lower']).copy()
      if not df_bolsas.empty:
        df_totales_b = pd.DataFrame([{
            "Mes Formateado": "TOTALES",
            "Unidad": "",
            "Cantidad": df_bolsas["Cantidad"].sum(),
            "N_Cotizaciones": df_bolsas["N_Cotizaciones"].sum(),
            "Pipeline_Bruto": df_bolsas["Pipeline_Bruto"].sum()
        }])
        df_bolsas = pd.concat([df_bolsas, df_totales_b], ignore_index=True)

      # --- 2. TABLA OTROS PRODUCTOS ---
      # Todo lo que NO sea bolsas ni servicio(s)
      df_otros = df_seg_base[~df_seg_base['Unidad_lower'].isin(['bolsas', 'servicio', 'servicios'])].drop(columns=['Unidad_lower']).copy()
      # No se agrega fila de totales por solicitud

      # --- 3. TABLA SERVICIOS ---
      df_servicios = df_seg_base[df_seg_base['Unidad_lower'].isin(['servicio', 'servicios'])].drop(columns=['Unidad_lower']).copy()
      if not df_servicios.empty:
        df_totales_s = pd.DataFrame([{
            "Mes Formateado": "TOTALES",
            "Unidad": "",
            "Cantidad": df_servicios["Cantidad"].sum(),
            "N_Cotizaciones": df_servicios["N_Cotizaciones"].sum(),
            "Pipeline_Bruto": df_servicios["Pipeline_Bruto"].sum()
        }])
        df_servicios = pd.concat([df_servicios, df_totales_s], ignore_index=True)

      # --- FUNCIÓN AUXILIAR DE FORMATEO ---
      def format_display_df(df):
        if df.empty: return df
        df_disp = df.copy()
        df_disp["Cantidad"] = df_disp["Cantidad"].apply(lambda x: f"{int(x):,}")
        df_disp["N_Cotizaciones"] = df_disp["N_Cotizaciones"].apply(lambda x: f"{int(x):,}")
        df_disp["Pipeline_Bruto"] = df_disp["Pipeline_Bruto"].apply(lambda x: f"S/ {x:,.2f}")
        return df_disp

      # --- MOSTRAR LAS TRES TABLAS ---
      st.markdown("#### 🛍️ Bolsas")
      if not df_bolsas.empty:
        st.dataframe(format_display_df(df_bolsas).style.set_properties(**{'text-align': 'left'}), use_container_width=True)
      else:
        st.info("No se encontraron registros de Bolsas.")

      st.markdown("#### 📦 Otros Productos")
      if not df_otros.empty:
        st.dataframe(format_display_df(df_otros).style.set_properties(**{'text-align': 'left'}), use_container_width=True)
      else:
        st.info("No se encontraron registros de Otros Productos.")

      st.markdown("#### 🔧 Servicios")
      if not df_servicios.empty:
        st.dataframe(format_display_df(df_servicios).style.set_properties(**{'text-align': 'left'}), use_container_width=True)
      else:
        st.info("No se encontraron registros de Servicios.")

    else:
      st.warning("Faltan columnas requeridas.")
  else:
    st.info("Sube tus archivos para generar el Seguimiento de Ventas.")

with tab3:
  st.subheader("Detalle General y Auditoría de Cotizaciones")
  if not st.session_state["df_auditoria"].empty:
    df_aud_disp = st.session_state["df_auditoria"].copy()
    
    # --- MODIFICACIÓN: Quitar columnas ponderadas SOLO de la visualización ---
    cols_a_quitar = ["Valor Ponderado (S/)", "Cantidad Ponderada Prod."]
    df_aud_disp = df_aud_disp.drop(columns=[c for c in cols_a_quitar if c in df_aud_disp.columns])
    
    # Buscador global para filtrar cualquier celda del historial (Cliente, N° Cotización, Vendedor, etc.)
    buscar_tab3 = st.text_input("🔍 Buscar cualquier dato (Cliente, Cotización, Vendedor, Unidad...):", "", key="b_tab3")
    
    if buscar_tab3:
      mask = df_aud_disp.astype(str).apply(lambda x: x.str.contains(buscar_tab3, case=False, na=False)).any(axis=1)
      df_aud_disp = df_aud_disp[mask]

    # --- NUEVO: Formateo de Columnas de Fecha a DD/MM/YYYY ---
    columnas_fechas = ["1o. Contacto", "Ultimo Contacto", "Mes Previsto"]
    for col in columnas_fechas:
      if col in df_aud_disp.columns:
        # Se convierte a fecha y se aplica el formato dd/mm/yyyy. Si el valor es inválido o vacío, se deja en blanco o se mantiene el original.
        df_aud_disp[col] = pd.to_datetime(df_aud_disp[col], errors='coerce').dt.strftime('%d/%m/%Y').fillna("")

    # Formateo de Contabilidad y Porcentajes
    if "Valor + IGV" in df_aud_disp.columns:
      df_aud_disp["Valor + IGV"] = df_aud_disp["Valor + IGV"].apply(lambda x: f"S/ {float(x):,.2f}" if pd.notna(x) and str(x).strip() != "" else "")
    if "Chance de Venta" in df_aud_disp.columns:
      df_aud_disp["Chance de Venta"] = df_aud_disp["Chance de Venta"].apply(lambda x: f"{float(x)*100:,.2f}%" if pd.notna(x) and str(x).strip() != "" else "")
    if "Cantidad" in df_aud_disp.columns:
      df_aud_disp["Cantidad"] = df_aud_disp["Cantidad"].apply(lambda x: f"{int(float(x)):,}" if pd.notna(x) and str(x).strip() != "" else "")
      
    st.dataframe(df_aud_disp.style.set_properties(**{'text-align': 'left'}), use_container_width=True)
    st.success(f"Registros encontrados / mostrados: {len(df_aud_disp)}")
  else:
    st.info("Sube tus archivos.")
